from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import load_workbook
from loguru import logger

MARGIN = 50
FONT_SIZE = 10
FONT_NAME = "Helvetica"


def convert(input_path: str, output_path: str) -> str:
    """Convert Excel file to PDF."""
    logger.info("Converting {} to PDF", input_path)

    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"File not found: {input_path}")

    wb = load_workbook(str(input_path), read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("No active worksheet found")

    data = []
    for row in ws.iter_rows(values_only=True):
        if any(cell is not None for cell in row):
            data.append([str(cell) if cell is not None else "" for cell in row])

    wb.close()

    if not data:
        raise ValueError("Worksheet is empty")

    width, height = A4
    available_width = width - 2 * MARGIN
    num_cols = len(data[0])
    col_widths = [available_width / num_cols] * num_cols

    table = Table(data, colWidths=col_widths)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, -1), FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), FONT_SIZE),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=MARGIN,
    )
    doc.build([table])

    logger.info("Saved PDF to: {}", output_path)
    return output_path
